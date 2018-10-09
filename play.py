import openpyxl

wb = openpyxl.load_workbook('cSpace_Bookingv1.xlsx')
# filepath='/home/code/cSpace_Bookingv1.xlsx'
# sheet=wb.active
# sheet = wb.get_sheet_by_name("Clients")
sheet = wb["Clients"]


nameslist = []


for cell in sheet['A']:

	#print("-->", cell)
	nameslist.append(cell.value)


# filepath='/home/code/cSpace_Bookingv1.xlsx'
# sheet=wb.active
# sheet = wb.get_sheet_by_name("Clients")
or cell in sheet['A']:
	#print("-->", cell)
	nameslist.append(cell.value)
# print(nameslist)

#received error when tried to use if statement.




# dict = {"first name": [], "last name": []}

# for cell in sheet['A']:
# 		first = row[0].split()[0]
# 		last = row[0].split()[1]
# 		dict["first name"].append(first)
# 		dict["last name"].append(last)
# 	#print("-->", cell)
# 	# nameslist.append(cell.value)
# Print(dict)

# print(nameslist)

# def nameslist(nam

# 	names_dict = {}
# 	for cell in sheet ['A']:
# 			if name in names_dict:
# 				name[k] = v
# 				first, last = name.split(' ')
# print(names_dict)
# make a list out of Column A (names)


# firstnames=([x.split()[0] for x in nameslist])
# print(firstnames)

# def occur_dict(a):
# 	my_dict = {}
# 	for i in range(21):
# 		cell_value_class = sh.cell(i,2).value
# 		cell_value_id = sh.cell(i,0).value
# 		my_dict[cell_value_class] = cell_value_id
# 	return my_dict
# 	print(my_dict)
#for rowOfCellObjects in sheet['A2':'F21']:
	#for cellObj in rowOfCellObjects:
		#print(cellObj.coordinate, cellObj.value)


#first_row = sheet.rows[1]
#print(first_row)
# a2 = sheet['A2']
# print(a2.value)
# names = sheet['A1' : 'A21']
# print(names)



# clients = {'name': 'fullname'}

# firstrest = names.partition(" ")
# for current_string in (firstrest):
    # print(current_string)





# namesissues=(names + issues)
# print(namesissues)


#names=sheet.cell('=A').value


#for row in sheet['A1':'A21']:
	#for cell in row:
		#print(cell.value)


#names=[]
#firstnames = [names.split(' ')[0] for name in nameslist]

#for rowOfCellObjects in sheet['A2':'A21']:
	#for cellObj in rowOfCellObjects:
			#print(cellObj.value)
#namesList = []
#for cell in sheet[]:
#namesList.append(cell.value)

#print(namesList)


#for row_data in new_row_data:
	#ws.append(row)

#for cell in sheet['A2:A21']:
	#nameslist.append





#full_name=cell

#print(cellObj)





