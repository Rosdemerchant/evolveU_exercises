import excel_names
import unittest
import openpyxl

wb = openpyxl.load_workbook('cSpace_Bookingv1.xlsx')
sheet = wb["Clients"]

class Test_practice(unittest.TestCase):
        
        def test_cell_has_value(self):
        	self.assertFalse(excel_names.cell_has_value(sheet, 1, 1))

        # ws.cell(row=1, column=1).value = ''
        # self.assertFalse(tdd_excel.cell_has_value(ws, 1, 1))

        # # ws.cell(row=1, column=1).value = '
        # # self.assertTrue(tdd_excel.cell_has_value(ws, 1, 1))

        # ws.cell(row=1, column=1).value = '   '
        # self.assertFalse(tdd_excel.cell_has_value(ws, 1, 1))

        # self.assertFalse(tdd_excel.cell_has_value(ws, 100, 100))
        # self.assertFalse(tdd_excel.cell_has_value(ws, 1, 100))
        # self.assertFalse(tdd_excel.cell_has_value(ws, 100, 1))
